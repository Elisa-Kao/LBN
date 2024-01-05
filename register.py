import openpyxl

def register(username, password):
    # Load the workbook
    wb = openpyxl.load_workbook('user_data.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Append a new row with the username and password
    sheet.append([username, password])

    # Save the changes
    wb.save('user_data.xlsx')

def login(username, password):
    # Load the workbook
    wb = openpyxl.load_workbook('user_data.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Check if the provided username and password match any row in the Excel file
    for row in sheet.iter_rows(values_only=True):
        if row[0] == username and row[1] == password:
            return True

    return False

# Example usage
register('user1', 'pass1')
result = login('user1', 'pass1')

if result:
    print('Login successful')
else:
    print('Login failed')
